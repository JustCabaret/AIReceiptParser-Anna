import io
import base64
import datetime
import json

from db import init_db, get_db_connection, get_receipt_details as fetch_receipt_details

# Dynamic imports for openpyxl
try:
    import openpyxl
except ImportError:
    openpyxl = None

# Dynamic imports for reportlab
try:
    import reportlab
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
except ImportError:
    reportlab = None

def export_receipts_excel(filters=None):
    """Generates Excel sheet (xlsx) for matching receipts and returns base64 string."""
    if openpyxl is None:
        raise ImportError(
            "Excel export requires openpyxl. Run: pip install -r executas/receipt-parser-tool/requirements.txt"
        )
    from openpyxl.styles import Font, Alignment, PatternFill
    
    init_db()
    connection = get_db_connection()
    try:
        cursor = connection.cursor()
        query = """
            SELECT 
                r.id as receipt_id, r.source as merchant, r.parsed_at as date,
                r.total as total_cents, r.warnings, r.tip as tip_cents, r.discount as discount_cents,
                ri.product as item_name, ri.quantity, ri.price as unit_price_cents, ri.category as item_category
            FROM receipts r
            LEFT JOIN receipt_items ri ON r.id = ri.receipt_id
        """
        
        conditions = []
        params = []
        
        if filters:
            if filters.get("receipt_ids"):
                placeholders = ",".join(["?"] * len(filters["receipt_ids"]))
                conditions.append(f"r.id IN ({placeholders})")
                params.extend(filters["receipt_ids"])
            if filters.get("date_from"):
                conditions.append("date(r.parsed_at) >= date(?)")
                params.append(filters["date_from"])
            if filters.get("date_to"):
                conditions.append("date(r.parsed_at) <= date(?)")
                params.append(filters["date_to"])
            if filters.get("category"):
                conditions.append("""
                    r.id IN (
                        SELECT receipt_id FROM receipt_items WHERE category LIKE ?
                    )
                """)
                params.append(filters["category"])
            if filters.get("merchant"):
                conditions.append("r.source LIKE ?")
                params.append(f"%{filters['merchant']}%")
                
        if conditions:
            query += " WHERE " + " AND ".join(conditions)
            
        query += " ORDER BY r.id DESC, ri.id ASC"
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Receipt Expenses"
        
        headers = [
            "Receipt ID", "Merchant", "Parsed Date", "Currency",
            "Subtotal", "Tax", "Tip", "Discount", "Total",
            "Item Name", "Quantity", "Unit Price", "Total Price", "Item Category",
            "Warnings"
        ]
        
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            
        for row_idx, row in enumerate(rows, 2):
            tip = int(row["tip_cents"] or 0)
            discount = int(row["discount_cents"] or 0)
            subtotal_c = int(row["total_cents"]) - tip + discount
            
            ws.cell(row=row_idx, column=1, value=row["receipt_id"])
            ws.cell(row=row_idx, column=2, value=row["merchant"])
            ws.cell(row=row_idx, column=3, value=row["date"])
            ws.cell(row=row_idx, column=4, value="EUR")
            
            ws.cell(row=row_idx, column=5, value=subtotal_c / 100.0).number_format = "#,##0.00"
            ws.cell(row=row_idx, column=6, value=0.0).number_format = "#,##0.00"
            ws.cell(row=row_idx, column=7, value=tip / 100.0).number_format = "#,##0.00"
            ws.cell(row=row_idx, column=8, value=discount / 100.0).number_format = "#,##0.00"
            ws.cell(row=row_idx, column=9, value=int(row["total_cents"]) / 100.0).number_format = "#,##0.00"
            
            ws.cell(row=row_idx, column=10, value=row["item_name"])
            ws.cell(row=row_idx, column=11, value=row["quantity"])
            ws.cell(row=row_idx, column=12, value=int(row["unit_price_cents"] or 0) / 100.0).number_format = "#,##0.00"
            ws.cell(row=row_idx, column=13, value=(int(row["unit_price_cents"] or 0) * int(row["quantity"] or 1)) / 100.0).number_format = "#,##0.00"
            ws.cell(row=row_idx, column=14, value=row["item_category"])
            
            warns_arr = []
            if row["warnings"]:
                try:
                    warns_arr = json.loads(row["warnings"])
                except Exception:
                    pass
            warns_str = ", ".join([w.get("message", "") for w in warns_arr]) if isinstance(warns_arr, list) else ""
            ws.cell(row=row_idx, column=15, value=warns_str)
            
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        out = io.BytesIO()
        wb.save(out)
        bytes_data = out.getvalue()
        
        if len(bytes_data) > 1.5 * 1024 * 1024:
            raise ValueError("Excel file size exceeds the 1.5 MB limit for stdio transport.")
            
        content_b64 = base64.b64encode(bytes_data).decode('utf-8')
        
        return {
            "success": True,
            "content_b64": content_b64,
            "filename": f"receipts_export_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx",
            "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
    finally:
        connection.close()

def export_receipts_pdf(filters=None):
    """Generates PDF report (summary list or single receipt ticket) and returns base64."""
    if reportlab is None:
        raise ImportError(
            "PDF export requires reportlab. Run: pip install -r executas/receipt-parser-tool/requirements.txt"
        )
    
    init_db()
    connection = get_db_connection()
    try:
        cursor = connection.cursor()
        
        mode = "summary"
        if filters and filters.get("mode"):
            mode = filters["mode"]
            
        query = """
            SELECT r.id, r.total, r.source as store, r.parsed_at, r.warnings, r.tip, r.discount
            FROM receipts r
        """
        conditions = []
        params = []
        
        if filters:
            if filters.get("receipt_ids"):
                placeholders = ",".join(["?"] * len(filters["receipt_ids"]))
                conditions.append(f"r.id IN ({placeholders})")
                params.extend(filters["receipt_ids"])
                if len(filters["receipt_ids"]) == 1:
                    mode = "receipt"
            if filters.get("date_from"):
                conditions.append("date(r.parsed_at) >= date(?)")
                params.append(filters["date_from"])
            if filters.get("date_to"):
                conditions.append("date(r.parsed_at) <= date(?)")
                params.append(filters["date_to"])
            if filters.get("category"):
                conditions.append("""
                    r.id IN (
                        SELECT receipt_id FROM receipt_items WHERE category LIKE ?
                    )
                """)
                params.append(filters["category"])
            if filters.get("merchant"):
                conditions.append("r.source LIKE ?")
                params.append(f"%{filters['merchant']}%")
                
        if conditions:
            query += " WHERE " + " AND ".join(conditions)
        query += " ORDER BY r.id DESC"
        
        cursor.execute(query, params)
        receipt_rows = cursor.fetchall()
        
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            name="TitleStyle",
            parent=styles["Heading1"],
            fontSize=18,
            leading=22,
            textColor=colors.HexColor("#3F51B5"),
            spaceAfter=15
        )
        subtitle_style = ParagraphStyle(
            name="SubTitleStyle",
            parent=styles["Normal"],
            fontSize=10,
            textColor=colors.gray,
            spaceAfter=20
        )
        body_style = styles["BodyText"]
        bold_style = ParagraphStyle(name="BoldStyle", parent=body_style, fontName="Helvetica-Bold")
        
        elements = []
        
        if mode == "receipt" and len(receipt_rows) > 0:
            row = receipt_rows[0]
            receipt_id = row["id"]
            
            details = fetch_receipt_details(receipt_id)
            
            elements.append(Paragraph(f"Receipt Expense Report (ID: {receipt_id})", title_style))
            elements.append(Paragraph(f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", subtitle_style))
            
            tip = details.get("tip") or 0
            discount = details.get("discount") or 0
            subtotal = details["total"] - tip + discount
            
            meta_data = [
                [Paragraph("<b>Merchant:</b>", body_style), Paragraph(details["store"], body_style)],
                [Paragraph("<b>Date:</b>", body_style), Paragraph(str(details["parsed_at"]), body_style)],
                [Paragraph("<b>Subtotal:</b>", body_style), Paragraph(f"EUR {subtotal/100:.2f}", body_style)],
                [Paragraph("<b>Tax:</b>", body_style), Paragraph("EUR 0.00", body_style)],
                [Paragraph("<b>Tip:</b>", body_style), Paragraph(f"EUR {tip/100:.2f}", body_style)],
                [Paragraph("<b>Discount:</b>", body_style), Paragraph(f"EUR {discount/100:.2f}", body_style)],
                [Paragraph("<b>Total:</b>", body_style), Paragraph(f"EUR {details['total']/100:.2f}", bold_style)],
            ]
            t_meta = Table(meta_data, colWidths=[100, 200])
            t_meta.setStyle(TableStyle([
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ]))
            elements.append(t_meta)
            elements.append(Spacer(1, 20))
            
            elements.append(Paragraph("<b>Items List</b>", styles["Heading3"]))
            elements.append(Spacer(1, 5))
            
            table_data = [["Product Name", "Quantity", "Unit Price", "Total Price", "Category"]]
            for item in details["items"]:
                p_name = item["name"]
                qty = item["quantity"]
                u_price = item["unit_price"]
                tot_price = u_price * qty
                cat = item["category"] or "Others"
                table_data.append([
                    Paragraph(p_name, body_style),
                    str(qty),
                    f"{u_price/100:.2f}",
                    f"{tot_price/100:.2f}",
                    cat
                ])
                
            t_items = Table(table_data, colWidths=[200, 50, 70, 70, 100])
            t_items.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#E0E0E0")),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('TOPPADDING', (0,0), (-1,-1), 6),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ]))
            elements.append(t_items)
            
            if details.get("warnings"):
                elements.append(Spacer(1, 20))
                elements.append(Paragraph("<b>Consistency Warnings</b>", styles["Heading3"]))
                for w in details["warnings"]:
                    elements.append(Paragraph(f"- [{w.get('code', 'WARNING')}] {w.get('message', '')}", body_style))
                    
        else:
            elements.append(Paragraph("Receipt Expense Summary Report", title_style))
            
            period_str = "All periods"
            if filters:
                df = filters.get("date_from")
                dt = filters.get("date_to")
                if df and dt:
                    period_str = f"Period: {df} to {dt}"
                elif df:
                    period_str = f"Period: From {df}"
                elif dt:
                    period_str = f"Period: Up to {dt}"
            elements.append(Paragraph(f"{period_str} | Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", subtitle_style))
            
            tot_scans = len(receipt_rows)
            tot_val = sum(r["total"] for r in receipt_rows)
            elements.append(Paragraph(f"<b>Total Spent:</b> EUR {tot_val/100:.2f} | <b>Total Scans:</b> {tot_scans}", body_style))
            elements.append(Spacer(1, 15))
            
            table_data = [["ID", "Date", "Merchant", "Tip", "Discount", "Total (EUR)"]]
            for r in receipt_rows:
                table_data.append([
                    str(r["id"]),
                    str(r["parsed_at"])[:10],
                    Paragraph(r["store"], body_style),
                    f"{r['tip']/100:.2f}",
                    f"{r['discount']/100:.2f}",
                    f"{r['total']/100:.2f}"
                ])
                
            t_receipts = Table(table_data, colWidths=[40, 70, 200, 60, 60, 80])
            t_receipts.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#3F51B5")),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('TOPPADDING', (0,0), (-1,-1), 6),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ]))
            elements.append(t_receipts)
            
        doc.build(elements)
        pdf_data = pdf_buffer.getvalue()
        
        if len(pdf_data) > 1.5 * 1024 * 1024:
            raise ValueError("PDF file size exceeds the 1.5 MB limit for stdio transport.")
            
        content_b64 = base64.b64encode(pdf_data).decode('utf-8')
        
        filename = f"receipt_report_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
        return {
            "success": True,
            "content_b64": content_b64,
            "filename": filename,
            "mime_type": "application/pdf"
        }
    finally:
        connection.close()
